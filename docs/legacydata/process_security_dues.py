#!/usr/bin/env python3
"""
Security Dues Processor for Residio
Processes legacy Excel payment tracker spreadsheets and generates structured JSON output.
"""

import json
import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime
from collections import defaultdict
from pathlib import Path
import re

# Column positions (1-indexed)
COL_HOUSE_NO = 1
COL_NAME = 2
COL_N = 3
COL_YEAR = 4
COL_RATE = 5
COL_JAN = 6
COL_DEC = 17
COL_PAID = 18
COL_DUE_DEBT = 19

MONTH_COLS = list(range(COL_JAN, COL_DEC + 1))
MONTH_NAMES = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

# Highlight color detection
YELLOW_THRESHOLD = (255, 255, 200)  # RGB for yellow highlights
BLUE_THRESHOLD = (150, 150, 255)     # RGB for blue highlights
RED_THRESHOLD = (255, 150, 150)      # RGB for red highlights

def is_color_match(fill, target_rgb, tolerance=50):
    """Check if a cell's fill color matches the target RGB within tolerance."""
    if not fill or not fill.fgColor:
        return False

    # Get RGB values
    rgb = fill.fgColor.rgb
    if not rgb or len(rgb) < 6:
        return False

    # Convert hex to RGB
    try:
        if isinstance(rgb, str):
            # Remove alpha if present (ARGB)
            if len(rgb) == 8:
                rgb = rgb[2:]
            r = int(rgb[0:2], 16)
            g = int(rgb[2:4], 16)
            b = int(rgb[4:6], 16)
        else:
            return False

        # Check if within tolerance
        return (abs(r - target_rgb[0]) <= tolerance and
                abs(g - target_rgb[1]) <= tolerance and
                abs(b - target_rgb[2]) <= tolerance)
    except:
        return False

def parse_currency(value):
    """Parse currency value, handling various formats."""
    if value is None or value == '':
        return 0.0

    if isinstance(value, (int, float)):
        return float(value)

    # Remove currency symbols, commas, spaces
    cleaned = str(value).replace('₦', '').replace(',', '').replace(' ', '').strip()

    # Handle parentheses (negative/credit)
    if cleaned.startswith('(') and cleaned.endswith(')'):
        cleaned = '-' + cleaned[1:-1]

    try:
        return float(cleaned) if cleaned else 0.0
    except:
        return 0.0

def extract_street_code(house_number):
    """Extract street code from house number (e.g., '19' -> '19', '20 F-1' -> '20')."""
    if not house_number:
        return ''

    # Take the first numeric part
    match = re.match(r'(\d+)', str(house_number).strip())
    return match.group(1) if match else str(house_number).strip()

def detect_property_type(name):
    """Detect if property is commercial based on name indicators."""
    if not name:
        return 'residential', []

    name_lower = str(name).lower()
    commercial_keywords = ['limited', 'ltd', 'plc', 'company', 'corp', 'business', 'enterprises']

    flags = []
    for keyword in commercial_keywords:
        if keyword in name_lower:
            flags.append('COMMERCIAL')
            return 'commercial', flags

    return 'residential', flags

def process_spreadsheet(file_path):
    """Process the Excel spreadsheet and extract payment data."""

    print(f"Loading spreadsheet: {file_path}")
    wb = openpyxl.load_workbook(file_path, data_only=False)
    ws = wb.active

    houses = {}
    current_house = None
    flags_summary = defaultdict(int)

    print(f"\nProcessing {ws.max_row} rows...")

    for row_idx in range(1, ws.max_row + 1):
        row = list(ws[row_idx])

        # Get house number
        house_no_cell = row[COL_HOUSE_NO - 1]
        house_no = house_no_cell.value

        # Skip empty rows
        if not house_no or str(house_no).strip() == '':
            continue

        house_no = str(house_no).strip()

        # Check if this is a new house block
        if house_no and house_no not in ['HOUSE NO', 'House No']:
            if house_no not in houses:
                # Initialize new house
                current_house = house_no
                houses[current_house] = {
                    'house_number': house_no,
                    'street_code': extract_street_code(house_no),
                    'primary_name': None,
                    'aliases': [],
                    'move_in_month': None,
                    'move_out_month': None,
                    'status': 'ACTIVE',
                    'years': [],
                    'flags': [],
                    'property_type': 'residential',
                    'rate_tier': None
                }
            else:
                current_house = house_no

        if not current_house:
            continue

        # Get resident name and check for yellow highlight (primary name)
        name_cell = row[COL_NAME - 1]
        name = name_cell.value

        if name and str(name).strip() and str(name).strip() not in ['NAME OF RESIDENT', 'Name of Resident']:
            name = str(name).strip()

            # Check for yellow highlight (primary name)
            if is_color_match(name_cell.fill, YELLOW_THRESHOLD):
                houses[current_house]['primary_name'] = name
            elif not houses[current_house]['primary_name']:
                # First name becomes primary if no yellow highlight detected
                houses[current_house]['primary_name'] = name
            elif name != houses[current_house]['primary_name'] and name not in houses[current_house]['aliases']:
                houses[current_house]['aliases'].append(name)

            # Check for commercial property
            prop_type, commercial_flags = detect_property_type(name)
            if commercial_flags:
                houses[current_house]['flags'].extend(commercial_flags)
                houses[current_house]['property_type'] = prop_type

        # Get year
        year_cell = row[COL_YEAR - 1]
        year = year_cell.value

        # Skip if not a valid year row
        if not year or not isinstance(year, (int, float)) or year < 2000 or year >= 2026:
            continue

        year = int(year)

        # Get rate
        rate_cell = row[COL_RATE - 1]
        rate = parse_currency(rate_cell.value)

        # Check for "OTHERS" rate tier
        if rate_cell.value and 'OTHERS' in str(rate_cell.value).upper():
            houses[current_house]['flags'].append('RATE_UNCLEAR')
            houses[current_house]['rate_tier'] = 'OTHERS'

        # Extract monthly payments
        payments = {}
        move_in_detected = None
        move_out_detected = None

        for i, month_col_idx in enumerate(MONTH_COLS):
            month_name = MONTH_NAMES[i]
            month_cell = row[month_col_idx - 1]
            amount = parse_currency(month_cell.value)

            payments[month_name] = amount

            # Check for blue highlight (move-in)
            if is_color_match(month_cell.fill, BLUE_THRESHOLD):
                move_in_detected = f"{year}-{i+1:02d}"

            # Check for red highlight (move-out)
            if is_color_match(month_cell.fill, RED_THRESHOLD):
                move_out_detected = f"{year}-{i+1:02d}"
                houses[current_house]['status'] = 'INACTIVE'

        # Update move-in/move-out if detected
        if move_in_detected and not houses[current_house]['move_in_month']:
            houses[current_house]['move_in_month'] = move_in_detected

        if move_out_detected:
            houses[current_house]['move_out_month'] = move_out_detected

        # Get totals
        paid_cell = row[COL_PAID - 1]
        paid_total = parse_currency(paid_cell.value)

        due_debt_cell = row[COL_DUE_DEBT - 1]
        due_debt = parse_currency(due_debt_cell.value)

        # Cross-validation: check if sum of monthly payments matches PAID column
        monthly_sum = sum(payments.values())
        variance = abs(monthly_sum - paid_total)

        year_flags = []
        if variance > 100:  # Allow ₦100 variance
            year_flags.append('SUM_MISMATCH')
            print(f"  Warning: House {current_house}, Year {year}: Sum mismatch (Σ={monthly_sum:,.2f}, PAID={paid_total:,.2f}, Δ={variance:,.2f})")

        # Calculate expected payment (rate * 12 for full year)
        expected = rate * 12
        year_balance = due_debt

        # Add year data
        houses[current_house]['years'].append({
            'year': year,
            'rate': rate,
            'payments': payments,
            'paid': paid_total,
            'expected': expected,
            'year_balance': year_balance,
            'flags': year_flags
        })

        # Update house flags
        for flag in year_flags:
            if flag not in houses[current_house]['flags']:
                houses[current_house]['flags'].append(flag)

    print(f"\nProcessed {len(houses)} house blocks")

    # Post-processing: Calculate net positions and finalize data
    for house_no, house_data in houses.items():
        # Calculate net position across all years
        total_expected = sum(year['expected'] for year in house_data['years'])
        total_paid = sum(year['paid'] for year in house_data['years'])
        net_position = total_paid - total_expected

        house_data['summary'] = {
            'total_expected': total_expected,
            'total_paid': total_paid,
            'net_position': net_position,
            'net_position_type': 'credit' if net_position > 0 else 'debt' if net_position < 0 else 'balanced',
            'currency': 'NGN'
        }

        # Set rate tier if not already set
        if not house_data['rate_tier'] and house_data['years']:
            # Use the most recent year's rate
            latest_rate = house_data['years'][-1]['rate']
            if latest_rate == 5000:
                house_data['rate_tier'] = 'TIER_1'
            elif latest_rate == 7000:
                house_data['rate_tier'] = 'TIER_2'
            elif latest_rate == 10000:
                house_data['rate_tier'] = 'TIER_3'
            else:
                house_data['rate_tier'] = 'OTHERS'

        # Ensure primary name exists
        if not house_data['primary_name']:
            house_data['flags'].append('NO_PRIMARY_NAME')
            house_data['primary_name'] = f"Resident at {house_no}"

    # Separate clean and flagged records
    clean_houses = []
    flagged_houses = []

    for house_no, house_data in houses.items():
        if house_data['flags']:
            flagged_houses.append(house_data)
            for flag in house_data['flags']:
                flags_summary[flag] += 1
        else:
            clean_houses.append(house_data)

    print(f"\nResults:")
    print(f"  Clean records: {len(clean_houses)}")
    print(f"  Flagged records: {len(flagged_houses)}")

    return clean_houses, flagged_houses, flags_summary, houses

def generate_summary(clean_houses, flagged_houses, flags_summary, all_houses, source_file):
    """Generate processing summary report."""

    total_houses = len(all_houses)
    total_residents = sum(1 + len(h.get('aliases', [])) for h in all_houses.values())

    # Financial summary
    total_expected = sum(h['summary']['total_expected'] for h in all_houses.values())
    total_paid = sum(h['summary']['total_paid'] for h in all_houses.values())
    net_position = total_paid - total_expected

    total_debt = sum(abs(h['summary']['net_position']) for h in all_houses.values()
                     if h['summary']['net_position'] < 0)
    total_credit = sum(h['summary']['net_position'] for h in all_houses.values()
                       if h['summary']['net_position'] > 0)

    # Data period
    all_years = set()
    for house in all_houses.values():
        for year_data in house['years']:
            all_years.add(year_data['year'])

    min_year = min(all_years) if all_years else None
    max_year = max(all_years) if all_years else None

    # Rate history
    rate_history = defaultdict(set)
    for house in all_houses.values():
        for year_data in house['years']:
            rate_history[year_data['year']].add(year_data['rate'])

    rate_history_formatted = {
        str(year): sorted(list(rates)) for year, rates in sorted(rate_history.items())
    }

    summary = {
        'export_metadata': {
            'export_date': datetime.now().isoformat(),
            'source_file': source_file,
            'interpretation_version': '1.0',
            'processor': 'Security Dues Processor for Residio'
        },
        'statistics': {
            'total_houses': total_houses,
            'clean_records': len(clean_houses),
            'flagged_records': len(flagged_houses),
            'total_residents': total_residents,
            'active_houses': sum(1 for h in all_houses.values() if h['status'] == 'ACTIVE'),
            'inactive_houses': sum(1 for h in all_houses.values() if h['status'] == 'INACTIVE')
        },
        'financial_summary': {
            'total_expected': total_expected,
            'total_paid': total_paid,
            'total_debt': total_debt,
            'total_credit': total_credit,
            'net_position': net_position,
            'net_position_type': 'credit' if net_position > 0 else 'debt' if net_position < 0 else 'balanced',
            'currency': 'NGN'
        },
        'data_period': {
            'start_year': min_year,
            'end_year': max_year,
            'years_covered': len(all_years),
            'note': 'Historical data only (years prior to 2026). Residio starts 2026 with calculated Net Position.'
        },
        'rate_history': rate_history_formatted,
        'flags_breakdown': dict(flags_summary),
        'validation_results': {
            'cross_validation_performed': True,
            'variance_threshold': 100,
            'highlight_detection_enabled': True
        }
    }

    return summary

def main():
    """Main processing function."""

    # File paths
    base_dir = Path(__file__).parent
    input_file = base_dir / 'ResidioTest.xlsx'
    output_dir = base_dir / 'importdata'

    # Ensure output directory exists
    output_dir.mkdir(exist_ok=True)

    # Process spreadsheet
    clean_houses, flagged_houses, flags_summary, all_houses = process_spreadsheet(input_file)

    # Generate output files
    print("\nGenerating output files...")

    # 1. Main import file (clean records)
    main_output = {
        'export_metadata': {
            'export_date': datetime.now().isoformat(),
            'source_file': str(input_file.name),
            'interpretation_version': '1.0',
            'total_houses': len(clean_houses),
            'data_period': {
                'start_year': min(year['year'] for h in clean_houses for year in h['years']) if clean_houses else None,
                'end_year': max(year['year'] for h in clean_houses for year in h['years']) if clean_houses else None
            }
        },
        'houses': clean_houses
    }

    main_file = output_dir / 'security_dues_import_main.json'
    with open(main_file, 'w') as f:
        json.dump(main_output, f, indent=2)
    print(f"  Created: {main_file}")

    # 2. Flagged records file
    flagged_output = {
        'export_metadata': {
            'export_date': datetime.now().isoformat(),
            'source_file': str(input_file.name),
            'interpretation_version': '1.0',
            'total_houses': len(flagged_houses),
            'note': 'These records require manual review before import'
        },
        'houses': flagged_houses
    }

    flagged_file = output_dir / 'security_dues_import_flagged.json'
    with open(flagged_file, 'w') as f:
        json.dump(flagged_output, f, indent=2)
    print(f"  Created: {flagged_file}")

    # 3. Summary report
    summary = generate_summary(clean_houses, flagged_houses, flags_summary, all_houses, str(input_file.name))

    summary_file = output_dir / 'security_dues_export_summary.json'
    with open(summary_file, 'w') as f:
        json.dump(summary, f, indent=2)
    print(f"  Created: {summary_file}")

    # Print summary
    print("\n" + "="*60)
    print("PROCESSING SUMMARY")
    print("="*60)
    print(f"\nTotal Houses Processed: {summary['statistics']['total_houses']}")
    print(f"  Clean Records: {summary['statistics']['clean_records']}")
    print(f"  Flagged Records: {summary['statistics']['flagged_records']}")
    print(f"  Total Residents: {summary['statistics']['total_residents']}")
    print(f"\nStatus:")
    print(f"  Active: {summary['statistics']['active_houses']}")
    print(f"  Inactive: {summary['statistics']['inactive_houses']}")
    print(f"\nFinancial Summary (NGN):")
    print(f"  Total Expected: ₦{summary['financial_summary']['total_expected']:,.2f}")
    print(f"  Total Paid: ₦{summary['financial_summary']['total_paid']:,.2f}")
    print(f"  Total Debt: ₦{summary['financial_summary']['total_debt']:,.2f}")
    print(f"  Total Credit: ₦{summary['financial_summary']['total_credit']:,.2f}")
    print(f"  Net Position: ₦{summary['financial_summary']['net_position']:,.2f} ({summary['financial_summary']['net_position_type']})")
    print(f"\nData Period: {summary['data_period']['start_year']} - {summary['data_period']['end_year']}")
    print(f"Years Covered: {summary['data_period']['years_covered']}")

    if flags_summary:
        print(f"\nFlags Breakdown:")
        for flag, count in sorted(flags_summary.items()):
            print(f"  {flag}: {count}")

    print("\n" + "="*60)
    print("Output files saved to:", output_dir)
    print("="*60)

if __name__ == '__main__':
    main()
