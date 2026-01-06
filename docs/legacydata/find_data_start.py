#!/usr/bin/env python3
"""Find where the actual data starts."""

import openpyxl

file_path = '/home/feyijimiohioma/projects/Residio/docs/legacydata/ResidioTest.xlsx'

print("Loading workbook...")
wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb.active

print(f"Searching for header row...")
for i in range(1, min(50, ws.max_row + 1)):
    row = ws[i]
    first_val = str(row[0].value).upper() if row[0].value else ""

    if 'HOUSE' in first_val and 'NO' in first_val:
        print(f"\nFound header at row {i}")
        print("Header row:")
        for j, cell in enumerate(row[:19], 1):
            print(f"  Col {j}: {cell.value}")

        print(f"\nFirst data row (row {i+1}):")
        next_row = ws[i+1]
        for j, cell in enumerate(next_row[:19], 1):
            print(f"  Col {j}: {cell.value}")
        break

print(f"\nChecking rows 15-25:")
for i in range(15, 26):
    row = ws[i]
    values = [str(cell.value)[:20] if cell.value else '' for cell in row[:5]]
    print(f"Row {i}: {values}")
