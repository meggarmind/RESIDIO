#!/usr/bin/env python3
"""Quick inspection of the Excel file structure."""

import openpyxl

file_path = '/home/feyijimiohioma/projects/Residio/docs/legacydata/ResidioTest.xlsx'

print("Loading workbook...")
wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb.active

print(f"Sheet name: {ws.title}")
print(f"Max row: {ws.max_row}")
print(f"Max column: {ws.max_column}")

print("\nFirst 10 rows:")
for i in range(1, min(11, ws.max_row + 1)):
    row = ws[i]
    values = [cell.value for cell in row[:19]]
    print(f"Row {i}: {values[:5]}...")  # Show first 5 columns

print("\nSample row values (Row 5):")
if ws.max_row >= 5:
    row = ws[5]
    for i, cell in enumerate(row[:19], 1):
        print(f"  Col {i}: {cell.value}")
