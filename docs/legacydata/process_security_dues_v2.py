#!/usr/bin/env python3
"""
Security Dues Processor for Residio - Version 2
Processes legacy Excel payment tracker spreadsheets and generates structured JSON output.
"""

import json
import openpyxl
from datetime import datetime
from collections import defaultdict
from pathlib import Path
import re

# Column positions (1-indexed) - ACTUAL STRUCTURE
COL_HOUSE_NO = 1
COL_STATUS = 2
COL_NAME = 3
COL_CONTACTS = 4
COL_YEAR = 5
COL_RATE = 6
COL_JAN = 7
COL_DEC = 18
COL_PAID = 19
# Note: DUE DEBT column doesn't appear to exist in this sheet

MONTH_COLS = list(range(COL_JAN, COL_DEC + 1))
MONTH_NAMES = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

# Start row for data (after header)
DATA_START_ROW = 16

def get_rgb_from_cell(cell):
    """Extract RGB values from cell fill."""
    if not cell.fill or not cell.fill.fgColor:
        return None
    rgb = cell.fill.fgColor.rgb
    if not rgb:
        return None
    try:
        # Handle different RGB formats
        if isinstance(rgb, str):
            if len(rgb) == 8:
                rgb = rgb[2:]  # Remove alpha (ARGB -> RGB)
            if len(rgb) >= 6:
                r = int(rgb[0:2], 16)
                g = int(rgb[2:4], 16)
                b = int(rgb[4:6], 16)
                return (r, g, b)
        # Handle RGB object
        elif hasattr(rgb, 'rgb'):
            return get_rgb_from_cell(cell)  # Recursively get string
        return None
    except:
        return None

def is_yellow_fill(cell):
    """Check if cell has yellow fill."""
    rgb = get_rgb_from_cell(cell)
    if not rgb:
        return False
    r, g, b = rgb
    # Yellow: high R and G, low B
    return r > 200 and g > 200 and b < 150

def is_blue_fill(cell):
    """Check if cell has blue fill."""
    rgb = get_rgb_from_cell(cell)
    if not rgb:
        return False
    r, g, b = rgb
    # Blue: low R and G, high B
    return b > 150 and r < 150 and g < 150

def is_red_fill(cell):
    """Check if cell has red fill."""
    rgb = get_rgb_from_cell(cell)
    if not rgb:
        return False
    r, g, b = rgb
    # Red: high R, low G and B
    return r > 200 and g < 150 and b < 150

def parse_currency(value):
    """Parse currency value, handling various formats."""
    if value is None or value == '':
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = str(value).replace('₦', '').replace(',', '').replace(' ', '').strip()
    if cleaned.startswith('(') and cleaned.endswith(')'):
        cleaned = '-' + cleaned[1:-1]
    try:
        return float(cleaned) if cleaned else 0.0
    except:
        return 0.0

def extract_street_code(house_number):
    """Extract street code from house number."""
    if not house_number:
        return ''
    match = re.match(r'(\d+)', str(house_number).strip())
    return match.group(1) if match else str(house_number).strip()

def process_spreadsheet(file_path):
    """Process the Excel spreadsheet and extract payment data."""

    print(f"Loading spreadsheet: {file_path}")
    wb = openpyxl.load_workbook(file_path, data_only=False)
    ws = wb.active

    houses = {}
    current_house = None
    row_count = 0

    print(f"\nProcessing from row {DATA_START_ROW} to {ws.max_row}...")

    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        row = list(ws[row_idx])

        # Get house number
        house_no_cell = row[COL_HOUSE_NO - 1]
        house_no_raw = house_no_cell.value

        # Check if this is a new house block
        if house_no_raw and str(house_no_raw).strip():
            house_no = str(house_no_raw).strip()

            # Skip if it looks like a header row
            if 'HOUSE' in house_no.upper():
                continue

            if house_no not in houses:
                current_house = house_no
                houses[current_house] = {
                    'house_number': house_no,
                    'street_code': extract_street_code(house_no),
                    'primary_name': None,
                    'aliases': [],
                    'move_in_month': None,
                    'move_out_month': None,
                    'status': 'ACTIVE',
                    'years': [],
                    'flags': [],
                    'property_type': 'residential',
                    'rate_tier': None
                }
                print(f"  Found house: {house_no}")
            else:
                current_house = house_no

        if not current_house:
            continue

        # Get resident name
        name_cell = row[COL_NAME - 1]
        name = name_cell.value

        if name and str(name).strip():
            name = str(name).strip()

            # Check for yellow highlight (primary name)
            if is_yellow_fill(name_cell):
                houses[current_house]['primary_name'] = name
            elif not houses[current_house]['primary_name']:
                houses[current_house]['primary_name'] = name
            elif name != houses[current_house]['primary_name'] and name not in houses[current_house]['aliases']:
                # Different name, add as alias
                if name not in houses[current_house]['aliases']:
                    houses[current_house]['aliases'].append(name)

        # Get year
        year_cell = row[COL_YEAR - 1]
        year = year_cell.value

        # Must have a valid year to process payment data
        if not year or not isinstance(year, (int, float)):
            continue

        year = int(year)

        # Skip future years (2026+)
        if year >= 2026:
            continue

        # Get rate
        rate_cell = row[COL_RATE - 1]
        rate = parse_currency(rate_cell.value)

        # Extract monthly payments
        payments = {}
        move_in_detected = None
        move_out_detected = None

        for i, month_col_idx in enumerate(MONTH_COLS):
            month_name = MONTH_NAMES[i]
            month_cell = row[month_col_idx - 1]
            amount = parse_currency(month_cell.value)

            payments[month_name] = amount

            # Check for blue highlight (move-in)
            if is_blue_fill(month_cell):
                move_in_detected = f"{year}-{i+1:02d}"

            # Check for red highlight (move-out)
            if is_red_fill(month_cell):
                move_out_detected = f"{year}-{i+1:02d}"
                houses[current_house]['status'] = 'INACTIVE'

        # Update move-in/move-out if detected
        if move_in_detected and not houses[current_house]['move_in_month']:
            houses[current_house]['move_in_month'] = move_in_detected

        if move_out_detected:
            houses[current_house]['move_out_month'] = move_out_detected

        # Get PAID total
        paid_cell = row[COL_PAID - 1]
        paid_total = parse_currency(paid_cell.value)

        # Cross-validation
        monthly_sum = sum(payments.values())
        variance = abs(monthly_sum - paid_total)

        year_flags = []
        if variance > 100:  # Allow ₦100 variance
            year_flags.append('SUM_MISMATCH')

        # Calculate expected (rate * 12 for full year)
        expected = rate * 12

        # Year balance is expected - paid
        year_balance = expected - paid_total

        # Add year data
        houses[current_house]['years'].append({
            'year': year,
            'rate': rate,
            'payments': payments,
            'paid': paid_total,
            'expected': expected,
            'year_balance': year_balance,
            'flags': year_flags
        })

        # Update house flags
        for flag in year_flags:
            if flag not in houses[current_house]['flags']:
                houses[current_house]['flags'].append(flag)

        row_count += 1

    print(f"\nProcessed {row_count} data rows")
    print(f"Found {len(houses)} house blocks")

    # Post-processing
    for house_no, house_data in houses.items():
        # Skip houses with no year data
        if not house_data['years']:
            continue

        # Calculate net position across all years
        total_expected = sum(year['expected'] for year in house_data['years'])
        total_paid = sum(year['paid'] for year in house_data['years'])
        net_position = total_paid - total_expected

        house_data['summary'] = {
            'total_expected': total_expected,
            'total_paid': total_paid,
            'net_position': net_position,
            'net_position_type': 'credit' if net_position > 0 else 'debt' if net_position < 0 else 'balanced',
            'currency': 'NGN'
        }

        # Set rate tier
        if not house_data['rate_tier'] and house_data['years']:
            latest_rate = house_data['years'][-1]['rate']
            if latest_rate == 5000:
                house_data['rate_tier'] = 'TIER_1'
            elif latest_rate == 7000:
                house_data['rate_tier'] = 'TIER_2'
            elif latest_rate == 10000:
                house_data['rate_tier'] = 'TIER_3'
            else:
                house_data['rate_tier'] = 'OTHERS'

        # Ensure primary name exists
        if not house_data['primary_name']:
            house_data['flags'].append('NO_PRIMARY_NAME')
            house_data['primary_name'] = f"Resident at {house_no}"

    # Remove houses with no years
    houses = {k: v for k, v in houses.items() if v['years']}

    # Separate clean and flagged records
    clean_houses = []
    flagged_houses = []
    flags_summary = defaultdict(int)

    for house_no, house_data in houses.items():
        if house_data['flags']:
            flagged_houses.append(house_data)
            for flag in house_data['flags']:
                flags_summary[flag] += 1
        else:
            clean_houses.append(house_data)

    print(f"\nResults:")
    print(f"  Clean records: {len(clean_houses)}")
    print(f"  Flagged records: {len(flagged_houses)}")

    return clean_houses, flagged_houses, flags_summary, houses

def generate_summary(clean_houses, flagged_houses, flags_summary, all_houses, source_file):
    """Generate processing summary report."""

    total_houses = len(all_houses)
    total_residents = sum(1 + len(h.get('aliases', [])) for h in all_houses.values())

    # Financial summary
    total_expected = sum(h['summary']['total_expected'] for h in all_houses.values())
    total_paid = sum(h['summary']['total_paid'] for h in all_houses.values())
    net_position = total_paid - total_expected

    total_debt = sum(abs(h['summary']['net_position']) for h in all_houses.values()
                     if h['summary']['net_position'] < 0)
    total_credit = sum(h['summary']['net_position'] for h in all_houses.values()
                       if h['summary']['net_position'] > 0)

    # Data period
    all_years = set()
    for house in all_houses.values():
        for year_data in house['years']:
            all_years.add(year_data['year'])

    min_year = min(all_years) if all_years else None
    max_year = max(all_years) if all_years else None

    # Rate history
    rate_history = defaultdict(set)
    for house in all_houses.values():
        for year_data in house['years']:
            rate_history[year_data['year']].add(year_data['rate'])

    rate_history_formatted = {
        str(year): sorted(list(rates)) for year, rates in sorted(rate_history.items())
    }

    summary = {
        'export_metadata': {
            'export_date': datetime.now().isoformat(),
            'source_file': source_file,
            'interpretation_version': '2.0',
            'processor': 'Security Dues Processor for Residio v2'
        },
        'statistics': {
            'total_houses': total_houses,
            'clean_records': len(clean_houses),
            'flagged_records': len(flagged_houses),
            'total_residents': total_residents,
            'active_houses': sum(1 for h in all_houses.values() if h['status'] == 'ACTIVE'),
            'inactive_houses': sum(1 for h in all_houses.values() if h['status'] == 'INACTIVE')
        },
        'financial_summary': {
            'total_expected': total_expected,
            'total_paid': total_paid,
            'total_debt': total_debt,
            'total_credit': total_credit,
            'net_position': net_position,
            'net_position_type': 'credit' if net_position > 0 else 'debt' if net_position < 0 else 'balanced',
            'currency': 'NGN'
        },
        'data_period': {
            'start_year': min_year,
            'end_year': max_year,
            'years_covered': len(all_years),
            'note': 'Historical data only (years prior to 2026). Residio starts 2026 with calculated Net Position.'
        },
        'rate_history': rate_history_formatted,
        'flags_breakdown': dict(flags_summary),
        'validation_results': {
            'cross_validation_performed': True,
            'variance_threshold': 100,
            'highlight_detection_enabled': True
        }
    }

    return summary

def main():
    """Main processing function."""

    base_dir = Path(__file__).parent
    input_file = base_dir / 'ResidioTest.xlsx'
    output_dir = base_dir / 'importdata'

    output_dir.mkdir(exist_ok=True)

    # Process spreadsheet
    clean_houses, flagged_houses, flags_summary, all_houses = process_spreadsheet(input_file)

    print("\nGenerating output files...")

    # 1. Main import file
    main_output = {
        'export_metadata': {
            'export_date': datetime.now().isoformat(),
            'source_file': str(input_file.name),
            'interpretation_version': '2.0',
            'total_houses': len(clean_houses),
            'data_period': {
                'start_year': min((year['year'] for h in clean_houses for year in h['years']), default=None),
                'end_year': max((year['year'] for h in clean_houses for year in h['years']), default=None)
            }
        },
        'houses': clean_houses
    }

    main_file = output_dir / 'security_dues_import_main.json'
    with open(main_file, 'w') as f:
        json.dump(main_output, f, indent=2)
    print(f"  Created: {main_file}")

    # 2. Flagged records file
    flagged_output = {
        'export_metadata': {
            'export_date': datetime.now().isoformat(),
            'source_file': str(input_file.name),
            'interpretation_version': '2.0',
            'total_houses': len(flagged_houses),
            'note': 'These records require manual review before import'
        },
        'houses': flagged_houses
    }

    flagged_file = output_dir / 'security_dues_import_flagged.json'
    with open(flagged_file, 'w') as f:
        json.dump(flagged_output, f, indent=2)
    print(f"  Created: {flagged_file}")

    # 3. Summary report
    summary = generate_summary(clean_houses, flagged_houses, flags_summary, all_houses, str(input_file.name))

    summary_file = output_dir / 'security_dues_export_summary.json'
    with open(summary_file, 'w') as f:
        json.dump(summary, f, indent=2)
    print(f"  Created: {summary_file}")

    # Print summary
    print("\n" + "="*60)
    print("PROCESSING SUMMARY")
    print("="*60)
    print(f"\nTotal Houses Processed: {summary['statistics']['total_houses']}")
    print(f"  Clean Records: {summary['statistics']['clean_records']}")
    print(f"  Flagged Records: {summary['statistics']['flagged_records']}")
    print(f"  Total Residents: {summary['statistics']['total_residents']}")
    print(f"\nStatus:")
    print(f"  Active: {summary['statistics']['active_houses']}")
    print(f"  Inactive: {summary['statistics']['inactive_houses']}")
    print(f"\nFinancial Summary (NGN):")
    print(f"  Total Expected: ₦{summary['financial_summary']['total_expected']:,.2f}")
    print(f"  Total Paid: ₦{summary['financial_summary']['total_paid']:,.2f}")
    print(f"  Total Debt: ₦{summary['financial_summary']['total_debt']:,.2f}")
    print(f"  Total Credit: ₦{summary['financial_summary']['total_credit']:,.2f}")
    print(f"  Net Position: ₦{summary['financial_summary']['net_position']:,.2f} ({summary['financial_summary']['net_position_type']})")
    print(f"\nData Period: {summary['data_period']['start_year']} - {summary['data_period']['end_year']}")
    print(f"Years Covered: {summary['data_period']['years_covered']}")

    if flags_summary:
        print(f"\nFlags Breakdown:")
        for flag, count in sorted(flags_summary.items()):
            print(f"  {flag}: {count}")

    print("\n" + "="*60)
    print("Output files saved to:", output_dir)
    print("="*60)

if __name__ == '__main__':
    main()
